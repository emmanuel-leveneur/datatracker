"""
Utilitaires pour l'import automatique de fichiers (CSV / Excel).
Détection d'encodage, de séparateur, inférence de types de colonnes.
"""
import csv
import io
import re
from datetime import date, datetime as dt
from typing import Any

import chardet
import openpyxl

from app.models import ColumnType

# ── Constantes ─────────────────────────────────────────────────────────────────

MAX_ROWS = 10_000
MAX_COLS = 50
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 Mo
SAMPLE_SIZE = 200  # nb de valeurs non vides analysées pour inférer le type

# Heuristique SELECT : au plus 15 valeurs distinctes ET moins de 20 % du total
SELECT_MAX_DISTINCT = 15
SELECT_MAX_RATIO = 0.20

# Patterns datetime supportés (AVANT les dates — ordre critique)
_DATETIME_PATTERNS = [
    (re.compile(r'^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}$'), '%Y-%m-%dT%H:%M:%S'),   # ISO T avec sec
    (re.compile(r'^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}$'),        '%Y-%m-%dT%H:%M'),      # ISO T sans sec
    (re.compile(r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$'),  '%Y-%m-%d %H:%M:%S'),  # ISO espace avec sec
    (re.compile(r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}$'),        '%Y-%m-%d %H:%M'),     # ISO espace sans sec
    (re.compile(r'^\d{2}/\d{2}/\d{4} \d{2}:\d{2}:\d{2}$'),  '%d/%m/%Y %H:%M:%S'),  # FR avec sec
    (re.compile(r'^\d{2}/\d{2}/\d{4} \d{2}:\d{2}$'),        '%d/%m/%Y %H:%M'),     # FR sans sec
]

# Patterns de dates supportés (ordre important : du plus précis au moins précis)
_DATE_PATTERNS = [
    (re.compile(r'^\d{4}-\d{2}-\d{2}$'), '%Y-%m-%d'),        # ISO
    (re.compile(r'^\d{2}/\d{2}/\d{4}$'), '%d/%m/%Y'),        # FR
    (re.compile(r'^\d{2}-\d{2}-\d{4}$'), '%d-%m-%Y'),        # FR tiret
    (re.compile(r'^\d{2}/\d{2}/\d{2}$'), '%d/%m/%y'),        # FR court
    (re.compile(r'^\d{4}/\d{2}/\d{2}$'), '%Y/%m/%d'),        # ISO slash
]

_BOOL_TRUE  = {'oui', 'yes', 'true', '1', 'vrai', 'o', 'y'}
_BOOL_FALSE = {'non', 'no', 'false', '0', 'faux', 'n'}
_BOOL_ALL   = _BOOL_TRUE | _BOOL_FALSE

_EMAIL_RE = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')

# Mots-clés de noms de colonnes géographiques (signal primaire pour LAT/LON)
_LATITUDE_KEYWORDS  = {"lat", "latitude"}
_LONGITUDE_KEYWORDS = {"lon", "lng", "long", "longitude"}


# ── Encodage & séparateur ──────────────────────────────────────────────────────

def detect_encoding(raw: bytes) -> str:
    result = chardet.detect(raw)
    enc = result.get('encoding') or 'utf-8'
    # Normaliser les alias Windows courants
    if enc.lower() in ('windows-1252', 'cp1252', 'ansi'):
        return 'cp1252'
    return enc


def detect_separator(text: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=',;\t|')
        return dialect.delimiter
    except csv.Error:
        return ';'


# ── Parsing ────────────────────────────────────────────────────────────────────

def parse_csv(raw: bytes) -> tuple[list[str], list[list[str]], str | None]:
    """
    Retourne (headers, rows, warning).
    rows est limité à MAX_ROWS. warning est un message si le fichier est tronqué.
    """
    encoding = detect_encoding(raw)
    try:
        text = raw.decode(encoding, errors='replace')
    except LookupError:
        text = raw.decode('utf-8', errors='replace')

    sep = detect_separator(text)
    reader = csv.reader(io.StringIO(text), delimiter=sep)

    rows_all = list(reader)
    if not rows_all:
        return [], [], None

    headers = [h.strip() for h in rows_all[0]]
    data_rows = [[cell.strip() for cell in row] for row in rows_all[1:] if any(c.strip() for c in row)]

    warning = None
    if len(data_rows) > MAX_ROWS:
        warning = f"Le fichier contient {len(data_rows):,} lignes — seules les {MAX_ROWS:,} premières ont été importées."
        data_rows = data_rows[:MAX_ROWS]

    return headers, data_rows, warning


def _cell_to_str(cell) -> str:
    """Convertit une cellule openpyxl en string sans perdre les entiers.
    openpyxl retourne les entiers Excel comme float Python (42 → 42.0) ;
    on les reconvertit en int avant de stringifier pour éviter le '42.0'.
    """
    if cell is None:
        return ''
    if isinstance(cell, float) and cell == int(cell):
        return str(int(cell))
    return str(cell).strip()


def parse_excel(raw: bytes, sheet_index: int = 0) -> tuple[list[str], list[list[str]], list[str], str | None]:
    """
    Retourne (headers, rows, sheet_names, warning).
    """
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    sheet_index = max(0, min(sheet_index, len(sheet_names) - 1))
    ws = wb.worksheets[sheet_index]

    rows_all = []
    for row in ws.iter_rows(values_only=True):
        rows_all.append([_cell_to_str(cell) for cell in row])

    if not rows_all:
        return [], [], sheet_names, None

    headers = [h.strip() for h in rows_all[0]]
    data_rows = [row for row in rows_all[1:] if any(c for c in row)]

    warning = None
    if len(data_rows) > MAX_ROWS:
        warning = f"Le fichier contient {len(data_rows):,} lignes — seules les {MAX_ROWS:,} premières ont été importées."
        data_rows = data_rows[:MAX_ROWS]

    return headers, data_rows, sheet_names, warning


# ── Inférence de types ─────────────────────────────────────────────────────────

def _is_datetime(val: str) -> bool:
    for pattern, fmt in _DATETIME_PATTERNS:
        if pattern.match(val):
            try:
                dt.strptime(val, fmt)
                return True
            except ValueError:
                pass
    return False


def _is_date(val: str) -> bool:
    for pattern, fmt in _DATE_PATTERNS:
        if pattern.match(val):
            try:
                dt.strptime(val, fmt)
                return True
            except ValueError:
                pass
    return False


def _is_bool_column(sample: list[str]) -> bool:
    """Retourne True si TOUTES les valeurs distinctes sont dans _BOOL_ALL (et au moins 2 distinctes)."""
    distinct = {v.lower().strip() for v in sample if v.strip()}
    return len(distinct) >= 2 and distinct.issubset(_BOOL_ALL)


def _is_integer(val: str) -> bool:
    # Rejette les valeurs avec zéro initial (ex: "007", "01234")
    if re.match(r'^-?0\d+$', val):
        return False
    try:
        int(val)
        return True
    except ValueError:
        return False


def _is_float(val: str) -> bool:
    # Gère la virgule française
    normalized = val.replace(',', '.', 1)
    # Rejette les entiers avec zéro initial (pas un float numérique significatif)
    if '.' not in normalized and re.match(r'^-?0\d+$', val):
        return False
    try:
        float(normalized)
        return True
    except ValueError:
        return False


def _is_email(val: str) -> bool:
    return bool(_EMAIL_RE.match(val))


def _is_lat_value(val: str) -> bool:
    try:
        return -90 <= float(val.replace(',', '.')) <= 90
    except ValueError:
        return False


def _is_lon_value(val: str) -> bool:
    try:
        return -180 <= float(val.replace(',', '.')) <= 180
    except ValueError:
        return False


def infer_column_type(all_values: list[str], col_name: str = "") -> ColumnType:
    """
    Infère le type d'une colonne à partir de toutes ses valeurs.
    - Utilise un échantillon de SAMPLE_SIZE valeurs non vides pour le type
    - Utilise toutes les valeurs pour l'heuristique SELECT (en dernier recours)
    """
    non_empty = [v for v in all_values if v.strip()]
    if not non_empty:
        return ColumnType.TEXT

    sample = non_empty[:SAMPLE_SIZE]
    total = len(sample)

    # Chaque type doit matcher sur au moins 95 % de l'échantillon
    threshold = 0.95

    # LATITUDE / LONGITUDE : le nom de colonne est le signal primaire
    if col_name:
        normalized = re.sub(r'[\s_\-]', '', col_name.lower())
        if any(kw in normalized for kw in _LATITUDE_KEYWORDS):
            if sum(_is_lat_value(v) for v in sample) / total >= threshold:
                return ColumnType.LATITUDE
        elif any(kw in normalized for kw in _LONGITUDE_KEYWORDS):
            if sum(_is_lon_value(v) for v in sample) / total >= threshold:
                return ColumnType.LONGITUDE

    # BOOLEAN : toutes les valeurs distinctes doivent être dans _BOOL_ALL
    if _is_bool_column(sample):
        return ColumnType.BOOLEAN

    # DATETIME avant DATE (les patterns datetime sont plus spécifiques)
    if sum(_is_datetime(v) for v in sample) / total >= threshold:
        return ColumnType.DATETIME

    if sum(_is_date(v) for v in sample) / total >= threshold:
        return ColumnType.DATE

    if sum(_is_email(v) for v in sample) / total >= threshold:
        return ColumnType.EMAIL

    if sum(_is_integer(v) for v in sample) / total >= threshold:
        return ColumnType.INTEGER

    if sum(_is_float(v) for v in sample) / total >= threshold:
        return ColumnType.FLOAT

    # Heuristique SELECT en dernier recours (aucun type fort détecté)
    distinct = set(v.strip() for v in all_values if v.strip())
    if (len(distinct) <= SELECT_MAX_DISTINCT
            and len(distinct) / len(non_empty) < SELECT_MAX_RATIO
            and len(non_empty) >= 5):
        return ColumnType.SELECT

    return ColumnType.TEXT


# ── Nettoyage des en-têtes ─────────────────────────────────────────────────────

def sanitize_headers(raw_headers: list[str]) -> list[str]:
    """Nettoie et dédoublonne les noms de colonnes."""
    seen: dict[str, int] = {}
    result = []
    for h in raw_headers[:MAX_COLS]:
        name = re.sub(r'\s+', ' ', h.strip()) or 'Colonne'
        if name in seen:
            seen[name] += 1
            name = f'{name}_{seen[name]}'
        else:
            seen[name] = 1
        result.append(name)
    return result


# ── Normalisation des valeurs ──────────────────────────────────────────────────

def normalize_value(val: str, col_type: ColumnType) -> str:
    """Convertit une valeur brute vers le format de stockage interne."""
    val = val.strip()
    if not val:
        return ''

    if col_type == ColumnType.BOOLEAN:
        return 'true' if val.lower() in _BOOL_TRUE else 'false'

    if col_type == ColumnType.DATETIME:
        for pattern, fmt in _DATETIME_PATTERNS:
            if pattern.match(val):
                try:
                    return dt.strptime(val, fmt).strftime('%Y-%m-%dT%H:%M')
                except ValueError:
                    pass
        return val  # stocké tel quel si non parseable

    if col_type == ColumnType.DATE:
        for pattern, fmt in _DATE_PATTERNS:
            if pattern.match(val):
                try:
                    return dt.strptime(val, fmt).strftime('%Y-%m-%d')
                except ValueError:
                    pass
        return val  # stocké tel quel si non parseable

    if col_type in (ColumnType.FLOAT, ColumnType.LATITUDE, ColumnType.LONGITUDE):
        return val.replace(',', '.', 1)

    return val
