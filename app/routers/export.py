import io
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.database import get_db
from app.dependencies import (
    can_access_table, get_current_user, get_table_or_404,
    get_visible_columns, is_table_owner,
)
from app.models import ActivityLog, DataTable, TableOwner, TablePermission, TableRow, User

templates = Jinja2Templates(directory="app/templates")

_PII_KEYWORDS = {
    "nom", "prénom", "prenom", "email", "courriel", "adresse",
    "téléphone", "telephone", "mobile", "portable", "naissance",
    "identifiant", "login", "ip", "genre", "sexe",
    "nationalité", "nationalite", "salaire", "revenu",
    "santé", "sante", "médical", "medical",
    "biométrique", "biometrique", "photo",
    "carte", "cin", "passeport", "rib", "iban",
    "numéro", "numero", "siret", "siren",
}

_BASE_LEGALE_LABELS = {
    "consent": ("La personne a donné son accord", "Art. 6.1.a — Consentement"),
    "contract": ("C'est nécessaire pour un contrat", "Art. 6.1.b — Exécution d'un contrat"),
    "legal": ("La loi nous y oblige", "Art. 6.1.c — Obligation légale"),
    "vital": ("Protection d'un intérêt vital", "Art. 6.1.d — Intérêt vital"),
    "public": ("Mission d'intérêt public", "Art. 6.1.e — Mission d'intérêt public"),
    "legit": ("Intérêt légitime de l'entreprise", "Art. 6.1.f — Intérêt légitime"),
}


def _is_pii(col_name: str, col_type: str) -> bool:
    if col_type == "email":
        return True
    lower = col_name.lower()
    return any(kw in lower for kw in _PII_KEYWORDS)

router = APIRouter(prefix="/tables", tags=["export"])


@router.get("/{table_id}/export/excel")
def export_excel(
    table: DataTable = Depends(get_table_or_404),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not can_access_table(table, user, db):
        raise HTTPException(status_code=403)

    visible_cols = get_visible_columns(table, user, db)
    visible_ids = {c.id for c in visible_cols}

    wb = Workbook()
    ws = wb.active
    ws.title = table.name[:31]  # Excel sheet name limit

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, col in enumerate(visible_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col.name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = max(15, len(col.name) + 4)

    # Data rows
    rows = (
        db.query(TableRow)
        .filter_by(table_id=table.id)
        .order_by(TableRow.created_at.asc())
        .all()
    )
    for row_idx, row in enumerate(rows, start=2):
        cells = {cv.column_id: cv.value for cv in row.cell_values if cv.column_id in visible_ids}
        for col_idx, col in enumerate(visible_cols, start=1):
            value = cells.get(col.id, "")
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-filter
    if visible_cols:
        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(visible_cols)).column_letter}1"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{table.name.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{table_id}/export/rgpd", response_class=HTMLResponse)
def rgpd_report(
    request: Request,
    table: DataTable = Depends(get_table_or_404),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not is_table_owner(table, user, db) and not user.is_admin:
        raise HTTPException(status_code=403, detail="Accès réservé aux propriétaires de la table.")

    row_count = db.query(TableRow).filter(
        TableRow.table_id == table.id,
        TableRow.deleted_at.is_(None),
    ).count()

    columns_with_pii = [
        {
            "name": col.name,
            "type": col.col_type.value,
            "required": col.required,
            "is_pii": col.is_personal_data or _is_pii(col.name, col.col_type.value),
            "is_pii_explicit": col.is_personal_data,
        }
        for col in table.columns
    ]
    pii_count = sum(1 for c in columns_with_pii if c["is_pii"])

    owners = (
        db.query(User)
        .join(TableOwner, TableOwner.user_id == User.id)
        .filter(TableOwner.table_id == table.id)
        .all()
    )
    granted = (
        db.query(User, TablePermission.level)
        .join(TablePermission, TablePermission.user_id == User.id)
        .filter(TablePermission.table_id == table.id)
        .all()
    )

    recent_logs = (
        db.query(ActivityLog)
        .filter(ActivityLog.table_id == table.id)
        .order_by(ActivityLog.timestamp.desc())
        .limit(20)
        .all()
    )

    return templates.TemplateResponse(
        request, "tables/rgpd_report.html",
        {
            "table": table,
            "user": user,
            "generated_at": datetime.utcnow(),
            "row_count": row_count,
            "columns_with_pii": columns_with_pii,
            "pii_count": pii_count,
            "owners": owners,
            "granted": granted,
            "recent_logs": recent_logs,
            "base_legale_label": _BASE_LEGALE_LABELS.get(table.rgpd_base_legale, ("", ""))[0],
        "base_legale_article": _BASE_LEGALE_LABELS.get(table.rgpd_base_legale, ("", ""))[1],
        },
    )
